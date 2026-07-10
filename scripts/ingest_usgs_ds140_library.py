#!/usr/bin/env python3
"""Build a compact 1861-1992 library from the official USGS DS140 catalog.

The output is split by commodity so GitHub Pages can lazy-load one workbook at
a time. Every numeric worksheet cell is retained with its year, source header,
unit, worksheet coordinate, and missing-value counts. Values are not
interpolated, combined across worksheets, or converted by this project.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import tempfile
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import urljoin, urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
CATALOG_URL = "https://www.usgs.gov/centers/national-minerals-information-center/historical-statistics-mineral-commodities-united"
DEFAULT_OUTPUT = ROOT / "data" / "usgs-ds140"
USER_AGENT = "critical-minerals-history-research/3.0 (official USGS data ingestion)"
START_YEAR = 1861
END_YEAR = 1992


def slug(value: str) -> str:
    return "-".join("".join(ch if ch.isalnum() else " " for ch in value.lower()).split())


def clean_text(value: object) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").split())


class CatalogParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[dict]] = []
        self.row: list[dict] | None = None
        self.cell: dict | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = dict(attrs)
        if tag == "tr":
            self.row = []
        elif tag in {"td", "th"} and self.row is not None:
            self.cell = {"text": "", "href": None}
            self.row.append(self.cell)
        elif tag == "a" and self.cell is not None:
            self.cell["href"] = attributes.get("href")

    def handle_data(self, data: str) -> None:
        if self.cell is not None:
            self.cell["text"] += data

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"}:
            self.cell = None
        elif tag == "tr" and self.row is not None:
            self.rows.append(self.row)
            self.row = None


class LinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        href = dict(attrs).get("href")
        if tag == "a" and href and ".xlsx" in href.lower():
            self.links.append(href)


def fetch(url: str, binary: bool = False) -> bytes | str:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=60) as response:
        payload = response.read()
        return payload if binary else payload.decode("utf-8", "replace")


def discover_catalog() -> list[dict]:
    parser = CatalogParser()
    parser.feed(str(fetch(CATALOG_URL)))
    rows: list[dict] = []
    for cells in parser.rows:
        if len(cells) < 3 or not cells[1].get("href"):
            continue
        href = str(cells[1]["href"])
        if "data-series-140" not in href and not href.lower().split("?")[0].endswith(".xlsx"):
            continue
        title = clean_text(cells[0]["text"])
        update_text = clean_text(cells[2]["text"])
        if not title or not update_text.isdigit():
            continue
        rows.append({
            "id": slug(title),
            "title": title,
            "catalog_href": urljoin(CATALOG_URL, href),
            "update_year": int(update_text),
        })
    if len(rows) < 80:
        raise RuntimeError(f"USGS catalog discovery returned only {len(rows)} commodity rows")
    return rows


def resolve_download(item: dict) -> dict:
    href = item["catalog_href"]
    if href.lower().split("?")[0].endswith(".xlsx"):
        return {**item, "source_url": CATALOG_URL, "download_url": href}
    parser = LinkParser()
    parser.feed(str(fetch(href)))
    if not parser.links:
        raise RuntimeError(f"No XLSX link found on {href}")
    return {**item, "source_url": href, "download_url": urljoin(href, parser.links[0])}


def workbook_unit(worksheet, header_row: int) -> str:
    for row in range(header_row - 1, 0, -1):
        text = clean_text(worksheet.cell(row, 1).value).strip("[]")
        if not text:
            continue
        lowered = text.lower()
        if "unless otherwise noted" in lowered:
            text = re.split("unless otherwise noted", text, flags=re.IGNORECASE)[0].rstrip(" ,")
        for prefix in ("All quantities are in ", "All quantities in ", "All values are in ", "All values in "):
            if text.lower().startswith(prefix.lower()):
                text = text[len(prefix):]
                break
        if any(token in lowered for token in ("quantity", "value", "metric ton", "kilogram", "carat", "cubic")):
            return text
    return "See the USGS workbook heading and embedded notes"


def measure_unit(label: str, source_unit: str) -> tuple[str, str]:
    lowered = label.lower().replace(" ", "")
    if "unitvalue" in lowered and "98$/" in lowered:
        return "1998 U.S. dollars per metric ton", "real-1998-dollars"
    if "unitvalue" in lowered and "$/" in lowered:
        return "current U.S. dollars per metric ton", "nominal"
    if "value($)" in lowered:
        return "current U.S. dollars", "nominal"
    return source_unit, "not-price"


def missing_state(value: object) -> str:
    if value is None or clean_text(value) == "":
        return "blank"
    text = clean_text(value).upper()
    if text in {"NA", "N/A", "NOT AVAILABLE"}:
        return "not_available"
    if text in {"W", "WITHHELD"}:
        return "withheld"
    return "other_nonnumeric"


def find_header(worksheet) -> tuple[int, int] | None:
    for row in range(1, min(worksheet.max_row, 20) + 1):
        for column in range(1, min(worksheet.max_column, 40) + 1):
            if clean_text(worksheet.cell(row, column).value).lower() == "year":
                return row, column
    return None


def extract_workbook(item: dict, workbook_path: Path, access_date: str) -> dict:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    series: list[dict] = []
    observation_count = 0
    all_years: set[int] = set()
    used_series_ids: set[str] = set()

    for worksheet in workbook.worksheets:
        header = find_header(worksheet)
        if not header:
            continue
        header_row, year_column = header
        source_unit = workbook_unit(worksheet, header_row)
        series_id = slug(worksheet.title) or "series"
        if series_id in used_series_ids:
            series_id = f"{series_id}-{len(used_series_ids) + 1}"
        used_series_ids.add(series_id)
        year_rows: list[tuple[int, int]] = []
        for row in range(header_row + 1, worksheet.max_row + 1):
            raw_year = worksheet.cell(row, year_column).value
            year = int(raw_year) if isinstance(raw_year, (int, float)) and not isinstance(raw_year, bool) and float(raw_year).is_integer() else None
            if year is not None and START_YEAR <= year <= END_YEAR:
                year_rows.append((row, year))
        if not year_rows:
            continue

        measures: list[dict] = []
        used_measure_ids: set[str] = set()
        for column in range(year_column + 1, worksheet.max_column + 1):
            label = clean_text(worksheet.cell(header_row, column).value)
            if not label:
                continue
            measure_id = slug(label) or f"column-{column}"
            if measure_id in used_measure_ids:
                measure_id = f"{measure_id}-column-{column}"
            used_measure_ids.add(measure_id)
            observations: list[list[int | float]] = []
            missing = {"blank": 0, "not_available": 0, "withheld": 0, "other_nonnumeric": 0}
            for row, year in year_rows:
                value = worksheet.cell(row, column).value
                if isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value)):
                    observations.append([year, value, row])
                else:
                    missing[missing_state(value)] += 1
            if not observations:
                continue
            unit, price_basis = measure_unit(label, source_unit)
            years = [int(row[0]) for row in observations]
            observation_count += len(observations)
            all_years.update(years)
            measures.append({
                "id": measure_id,
                "label": label,
                "column": column,
                "unit": unit,
                "price_basis": price_basis,
                "observation_count": len(observations),
                "year_start": min(years),
                "year_end": max(years),
                "missing": missing,
                "observations": observations,
            })
        if measures:
            series.append({
                "id": series_id,
                "label": clean_text(worksheet.title),
                "worksheet": worksheet.title,
                "header_row": header_row,
                "source_unit": source_unit,
                "year_start": min(year for _, year in year_rows),
                "year_end": max(year for _, year in year_rows),
                "measures": measures,
            })

    return {
        "schema_version": 1,
        "commodity": {
            "id": item["id"],
            "title": item["title"],
            "update_year": item["update_year"],
            "source_url": item["source_url"],
            "download_url": item["download_url"],
            "catalog_url": CATALOG_URL,
            "access_date": access_date,
            "historical_scope": {"start": START_YEAR, "end": END_YEAR},
            "transcription_status": "machine-extracted-xlsx",
            "conversion_methodology": "No project conversion. Numeric values are reproduced from USGS-standardized workbook cells.",
            "caveat": "Blank, unavailable, withheld, and other nonnumeric cells are counted separately and are never treated as zero.",
        },
        "summary": {
            "series_count": len(series),
            "measure_count": sum(len(row["measures"]) for row in series),
            "observation_count": observation_count,
            "year_start": min(all_years) if all_years else None,
            "year_end": max(all_years) if all_years else None,
        },
        "series": series,
    }


def build_item(item: dict, cache_dir: Path, output_dir: Path, access_date: str) -> dict:
    resolved = resolve_download(item)
    filename = Path(urlparse(resolved["download_url"]).path).name
    workbook_path = cache_dir / filename
    if not workbook_path.exists():
        workbook_path.write_bytes(bytes(fetch(resolved["download_url"], binary=True)))
    payload = extract_workbook(resolved, workbook_path, access_date)
    output_path = output_dir / "commodities" / f"{resolved['id']}.json"
    output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return {
        **resolved,
        **payload["summary"],
        "data_url": f"data/usgs-ds140/commodities/{resolved['id']}.json",
        "status": "verified-numeric-extraction" if payload["summary"]["observation_count"] else "review-required",
    }


def run(output_dir: Path, cache_dir: Path, access_date: str, workers: int) -> dict:
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "commodities").mkdir(parents=True, exist_ok=True)
    cache_dir.mkdir(parents=True, exist_ok=True)
    discovered = discover_catalog()
    results: list[dict] = []
    failures: list[dict] = []
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(build_item, item, cache_dir, output_dir, access_date): item for item in discovered}
        for future in as_completed(futures):
            item = futures[future]
            try:
                results.append(future.result())
            except Exception as error:
                failures.append({**item, "status": "review-required", "error": str(error)})
    results.sort(key=lambda row: row["title"].lower())
    failures.sort(key=lambda row: row["title"].lower())
    manifest = {
        "schema_version": 1,
        "title": "USGS Data Series 140 historical commodity library",
        "catalog_url": CATALOG_URL,
        "access_date": access_date,
        "historical_scope": {"start": START_YEAR, "end": END_YEAR},
        "method": "Numeric XLSX cells only; no project interpolation, aggregation, or unit conversion.",
        "commodity_count": len(results),
        "series_count": sum(row["series_count"] for row in results),
        "measure_count": sum(row["measure_count"] for row in results),
        "observation_count": sum(row["observation_count"] for row in results),
        "review_queue_count": len(failures),
        "commodities": results + failures,
    }
    (output_dir / "catalog.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--cache-dir", type=Path)
    parser.add_argument("--access-date", default=date.today().isoformat())
    parser.add_argument("--workers", type=int, default=8)
    args = parser.parse_args()
    if args.cache_dir:
        manifest = run(args.output_dir, args.cache_dir, args.access_date, args.workers)
    else:
        with tempfile.TemporaryDirectory(prefix="usgs-ds140-library-") as directory:
            manifest = run(args.output_dir, Path(directory), args.access_date, args.workers)
    print(
        f"Wrote {manifest['observation_count']:,} observations across "
        f"{manifest['commodity_count']} commodities and {manifest['series_count']} worksheet series; "
        f"review queue={manifest['review_queue_count']}"
    )


if __name__ == "__main__":
    main()
